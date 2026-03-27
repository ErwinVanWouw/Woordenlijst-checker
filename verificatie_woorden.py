"""
verificatie_woorden.py — Automatische verificatie van Woordenlijst-checker
Vergelijkt de app-output per woord met handmatig geverifieerde verwachtingen
uit 'woordverificatie woordenlijst.org.xlsx'.

Gebruik:
    python verificatie_woorden.py

Vereist: openpyxl  (pip install openpyxl)

Structuur van het Excel-bestand:
    Rij 1 : volledige naam van elke woordsoort
    Rij 2 : afkortingen zoals de app ze toont (bijv. 'znw.', 'ww.', 'pers. vnw.')
    Rij 3+ : één woord per rij; cijfer in een cel = verwacht aantal entries
             voor die woordsoort; laatste kolom is totaal (SUM-formule, genegeerd)
"""

import os
import sys

try:
    import openpyxl
except ImportError:
    print("[Fout] openpyxl is niet geïnstalleerd.")
    print("       Voer uit: pip install openpyxl")
    sys.exit(1)

# Importeer gedeelde logica uit test_woorden.py (synchroon houden met woordenlijstchecker.py)
from test_woorden import check_word_online, normaliseer_apostrof

# ---------------------------------------------------------------------------
# Configuratie
# ---------------------------------------------------------------------------

EXCEL_PAD = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "woordverificatie woordenlijst.org.xlsx")

# Normaliseer Excel-afkortingen naar de schrijfwijze van de app
ABBREV_NORMALISATIE = {
    'bnw. / bw.': 'bnw./bw.',  # Excel gebruikt spaties rond slash, app niet
    'betr. vnw':  'betr. vnw.',  # Excel mist de afsluitende punt
}

# ---------------------------------------------------------------------------
# Excel inlezen
# ---------------------------------------------------------------------------

def laad_verwachtingen(pad):
    """
    Laad verwachte resultaten uit de Excel.
    Geeft dict terug: term → {'labels': {afkorting: aantal}, 'totaal': int}
    """
    wb = openpyxl.load_workbook(pad)
    ws = wb.active

    # Rij 2: afkortingen per kolom
    rij2 = [cel.value for cel in ws[2]]
    col_abbrev = {}
    for i, abbrev in enumerate(rij2):
        if abbrev and isinstance(abbrev, str):
            norm = ABBREV_NORMALISATIE.get(abbrev.strip(), abbrev.strip())
            col_abbrev[i] = norm

    verwachtingen = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        term = row[0]
        if term is None:
            continue
        term = str(term).strip()

        labels = {}
        totaal = 0
        for i, val in enumerate(row):
            if i == 0:
                continue  # term-kolom overslaan
            if val and isinstance(val, (int, float)) and int(val) > 0:
                abbrev = col_abbrev.get(i)
                if abbrev:
                    labels[abbrev] = labels.get(abbrev, 0) + int(val)
                    totaal += int(val)

        verwachtingen[term] = {'labels': labels, 'totaal': totaal}

    return verwachtingen

# ---------------------------------------------------------------------------
# Verificatielogica
# ---------------------------------------------------------------------------

def effectieve_label(entry):
    """Bereken het weergavelabel zoals de popup dat toont."""
    if entry.get('is_meervoud'):
        return 'znw. mv.'
    return entry.get('display', '')


def controleer_woord(woord, verwacht):
    """
    Controleer één woord tegen de verwachting.
    Geeft (geslaagd: bool, melding: str) terug.
    """
    woord_norm = normaliseer_apostrof(woord)
    is_valid, _, error_msg, _, word_info, _, _ = check_word_online(woord_norm)

    entries = (word_info or {}).get('entries', [])
    gevonden_labels = {}
    for entry in entries:
        lbl = effectieve_label(entry)
        if lbl:
            gevonden_labels[lbl] = gevonden_labels.get(lbl, 0) + 1
    gevonden_totaal = sum(gevonden_labels.values())

    verwacht_totaal = verwacht['totaal']
    verwacht_labels = verwacht['labels']

    if verwacht_totaal == 0:
        if not is_valid:
            reden = f"({error_msg})" if error_msg else ""
            return True, f"✓  niet gevonden {reden}".strip()
        else:
            return False, (
                f"✗  verwacht: NIET gevonden\n"
                f"       gevonden: {gevonden_labels} (totaal {gevonden_totaal})"
            )
    else:
        if not is_valid:
            return False, (
                f"✗  verwacht: {verwacht_labels} (totaal {verwacht_totaal})\n"
                f"       niet gevonden: {error_msg or '—'}"
            )
        totaal_ok = (gevonden_totaal == verwacht_totaal)
        labels_ok = (gevonden_labels == verwacht_labels)
        if totaal_ok and labels_ok:
            return True, f"✓  {gevonden_labels} (totaal {gevonden_totaal})"
        else:
            return False, (
                f"✗  verwacht: {verwacht_labels} (totaal {verwacht_totaal})\n"
                f"       gevonden: {gevonden_labels} (totaal {gevonden_totaal})"
            )

# ---------------------------------------------------------------------------
# Hoofdprogramma
# ---------------------------------------------------------------------------

def main():
    if not os.path.exists(EXCEL_PAD):
        print(f"[Fout] Excel niet gevonden: {EXCEL_PAD}")
        sys.exit(1)

    verwachtingen = laad_verwachtingen(EXCEL_PAD)

    BREED = 72
    print("=" * BREED)
    print("  Woordenlijst-checker — verificatierun")
    print(f"  Bron: {os.path.basename(EXCEL_PAD)}")
    print(f"  {len(verwachtingen)} woorden te controleren")
    print("=" * BREED)

    geslaagd = 0
    mislukt = 0
    mislukt_woorden = []

    for woord, verwacht in verwachtingen.items():
        woord_norm = normaliseer_apostrof(woord)
        norm_str = f" → '{woord_norm}'" if woord_norm != woord else ""
        if verwacht['totaal'] > 0:
            verwacht_str = f"[verwacht: {verwacht['totaal']}]"
        else:
            verwacht_str = "[verwacht: niet gevonden]"

        ok, melding = controleer_woord(woord, verwacht)

        print(f"\n{'─' * BREED}")
        print(f"  '{woord}'{norm_str}  {verwacht_str}")
        print(f"  {melding}")

        if ok:
            geslaagd += 1
        else:
            mislukt += 1
            mislukt_woorden.append(woord)

    print(f"\n{'=' * BREED}")
    print(f"  Resultaat: {geslaagd} geslaagd, {mislukt} mislukt "
          f"van {geslaagd + mislukt}")
    if mislukt_woorden:
        print(f"  Mislukt:   {', '.join(repr(w) for w in mislukt_woorden)}")
    print("=" * BREED)


if __name__ == "__main__":
    main()
